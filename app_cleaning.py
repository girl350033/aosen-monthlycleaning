import streamlit as st
import pandas as pd
import datetime
import calendar
import holidays
import io
import html

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from supabase import create_client


# =========================================================
# 頁面設定
# =========================================================

st.set_page_config(
    page_title="澳森托嬰中心 月清潔輪值表",
    page_icon="🏫",
    layout="wide",
)


# =========================================================
# CSS
# =========================================================

st.markdown(
    """
<style>

.block-container {
    padding-top: 0.7rem !important;
    padding-bottom: 0.8rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}

h1 {
    margin-top: 0 !important;
    margin-bottom: 0.45rem !important;
}

h2, h3, h4 {
    margin-top: 0.25rem !important;
    margin-bottom: 0.35rem !important;
}

hr {
    margin-top: 0.55rem !important;
    margin-bottom: 0.55rem !important;
}

div[data-testid="stHorizontalBlock"] {
    gap: 0.55rem !important;
}

div[data-testid="stVerticalBlock"] {
    gap: 0.45rem !important;
}

div[data-testid="stExpander"] {
    margin-bottom: 0.55rem !important;
}

div[data-testid="stExpander"] details {
    border-radius: 9px !important;
}

div[data-testid="stExpander"] summary {
    padding-top: 0.4rem !important;
    padding-bottom: 0.4rem !important;
    font-weight: 600 !important;
}

/* Selectbox */
div[data-baseweb="select"] > div {
    min-height: 56px !important;
    height: auto !important;
    border-radius: 9px !important;
}

div[data-baseweb="select"] span {
    white-space: normal !important;
    overflow: hidden !important;
    text-overflow: unset !important;
    line-height: 1.28 !important;

    display: -webkit-box !important;
    -webkit-line-clamp: 2 !important;
    -webkit-box-orient: vertical !important;

    max-height: 2.7em !important;
}

li[role="option"] {
    white-space: normal !important;
    height: auto !important;
    min-height: 40px !important;
    line-height: 1.3 !important;
}

/* 星期＋日期 */
.day-header {
    display: flex;
    align-items: center;
    gap: 6px;
    margin-top: 0;
    margin-bottom: 3px;
    min-height: 27px;
}

.day-name {
    font-size: 17px;
    font-weight: 700;
    white-space: nowrap;
}

.day-date {
    font-size: 13px;
    color: #12833b;
    background-color: #f4f6f5;
    padding: 2px 6px;
    border-radius: 6px;
    white-space: nowrap;
}

/* 預覽 */
.cleaning-preview-wrapper {
    width: 100%;
    max-width: 100%;
    overflow: hidden;
    margin-top: 4px;
}

.cleaning-preview-table {
    width: 100%;
    max-width: 100%;
    table-layout: fixed;
    border-collapse: collapse;
    font-size: 13px;
    background-color: white;
}

.cleaning-preview-table th {
    background-color: #f5f6f8;
    border: 1px solid #d9dce1;
    padding: 6px 4px;
    text-align: center;
    vertical-align: middle;
    font-weight: 700;
}

.cleaning-preview-table td {
    border: 1px solid #d9dce1;
    padding: 6px 5px;
    vertical-align: top;

    white-space: normal !important;
    word-break: break-word !important;
    overflow-wrap: anywhere !important;
}

.cleaning-preview-table .week-col {
    width: 6%;
    text-align: center;
    vertical-align: middle;
    font-weight: 700;
}

.cleaning-preview-table .day-col {
    width: 18.8%;
}

.preview-date {
    font-weight: 700;
    color: #12833b;
    margin-bottom: 3px;
}

.preview-task {
    font-size: 13px;
    line-height: 1.35;
    margin-bottom: 4px;
}

.preview-teacher {
    font-size: 12px;
    color: #666666;
}

/* 管理區 */
.manage-title {
    font-weight: 700;
    margin-bottom: 5px;
}

/* 小螢幕 */
@media (max-width: 1400px) {

    .block-container {
        padding-left: 0.6rem !important;
        padding-right: 0.6rem !important;
    }

    div[data-testid="stHorizontalBlock"] {
        gap: 0.35rem !important;
    }

    .day-name {
        font-size: 15px;
    }

    .day-date {
        font-size: 12px;
        padding: 2px 4px;
    }

    .cleaning-preview-table {
        font-size: 11px;
    }

    .preview-task {
        font-size: 11px;
    }

    .preview-teacher {
        font-size: 10px;
    }

    .cleaning-preview-table td {
        padding: 4px 3px;
    }
}

div[data-testid="stTabs"] button {
    padding-top: 0.45rem !important;
    padding-bottom: 0.45rem !important;
}

</style>
""",
    unsafe_allow_html=True,
)


# =========================================================
# Supabase 連線
# =========================================================

@st.cache_resource
def get_supabase():

    try:
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]

        return create_client(url, key)

    except Exception as error:

        st.error(
            "❌ 無法連線 Supabase。請確認 Streamlit Secrets "
            "中的 SUPABASE_URL 與 SUPABASE_KEY。"
        )

        st.exception(error)
        st.stop()


supabase = get_supabase()


# =========================================================
# 預設資料
# =========================================================

DEFAULT_TEACHERS_A = [
    "主任",
    "均宜",
    "小安",
    "綺綺",
    "嘉鳳",
    "樺樺",
    "Candy",
    "Panda",
]

DEFAULT_TEACHERS_B = [
    "主任",
    "均宜",
    "小安",
    "綺綺",
    "嘉鳳",
    "樺樺",
    "Candy",
    "Panda",
]


DEFAULT_TASKS_A = [
    "樓下酵素熱水倒水槽(幼兒洗手水槽、門口小水槽、洗屁屁區、備餐區，使用2/3桶熱水配半杓酵素攪拌均勻)",
    "樓梯間掃地拖地(包含牆壁檯面灰塵清潔)",
    "晨檢區灰塵及洗手台(檯面及角落灰塵清潔、洗手台用漂白水噴式擦拭、掃地及拖地)",
    "樓上廁所(洗手台漂白水噴拭除黴、馬桶清潔刷淨、地板掃地拖地)",
    "辦公室灰塵(檯面整理、灰塵擦拭乾淨、後方櫃子除塵、監視器及縫隙除塵)",
    "清點備品",
    "除濕機濾網跟接水槽清潔",
    "刪除上個月Line相簿",
    "廚房門口上面紗窗清潔(使用雞毛撢子撥灰塵)",
    "掃地機器人/寢具表填寫/規劃戶外活動計畫表與回報(髒水倒掉、洗淨、補充乾淨水、清潔液1:50)",
]


DEFAULT_TASKS_B = [
    "酵素熱水倒水槽(幼兒洗手水槽、門口小水槽、洗屁屁區、備餐區，使用2/3桶熱水配半杓酵素攪拌均勻)",
    "樓梯間掃地拖地(包含玻璃檯面灰塵清潔)",
    "晨檢區灰塵及洗手台(檯面及角落灰塵清潔、洗手台用漂白水噴式擦拭、掃地及拖地)",
    "樓上廁所(洗手台漂白水噴拭除黴、馬桶清潔刷淨、地板掃地拖地)",
    "辦公室灰塵(檯面整理、灰塵擦拭乾淨、後方櫃子除塵、監視器及縫隙除塵)",
    "清點備品",
    "除濕機濾網跟接水槽清潔",
    "刪除上個月Line相簿",
    "廚房門口上面紗窗清潔(使用雞毛撢子撥灰塵)",
    "戶外掃落葉/寢具表填寫/規劃戶外活動計畫表與回報",
]


DEFAULT_NOTES = (
    "備註：請各位老師確實執行清潔項目，"
    "若遇請假請務必提前找職務代理人協助。"
)


# =========================================================
# Supabase 資料庫函式
# =========================================================

def load_teachers(branch):

    response = (
        supabase
        .table("daycare_teachers")
        .select("*")
        .eq("branch", branch)
        .order("sort_order")
        .execute()
    )

    return response.data or []


def load_tasks(branch):

    response = (
        supabase
        .table("daycare_cleaning_tasks")
        .select("*")
        .eq("branch", branch)
        .order("sort_order")
        .execute()
    )

    return response.data or []


def load_settings(branch):

    response = (
        supabase
        .table("daycare_settings")
        .select("*")
        .eq("branch", branch)
        .execute()
    )

    data = response.data or []

    return data[0] if data else None


def load_history(branch):

    response = (
        supabase
        .table("daycare_schedule_history")
        .select("*")
        .eq("branch", branch)
        .order("created_at", desc=True)
        .execute()
    )

    return response.data or []


# =========================================================
# 第一次啟動：初始化預設資料
# =========================================================

def initialize_branch_data(
    branch,
    default_teachers,
    default_tasks,
):

    # -----------------------------------------------------
    # 老師
    # -----------------------------------------------------

    teachers = load_teachers(branch)

    if not teachers:

        rows = []

        for index, teacher in enumerate(default_teachers):

            rows.append(
                {
                    "branch": branch,
                    "name": teacher,
                    "sort_order": index,
                }
            )

        if rows:

            (
                supabase
                .table("daycare_teachers")
                .insert(rows)
                .execute()
            )

    # -----------------------------------------------------
    # 清潔項目
    # -----------------------------------------------------

    tasks = load_tasks(branch)

    if not tasks:

        rows = []

        for index, task in enumerate(default_tasks):

            rows.append(
                {
                    "branch": branch,
                    "task_name": task,
                    "sort_order": index,
                }
            )

        if rows:

            (
                supabase
                .table("daycare_cleaning_tasks")
                .insert(rows)
                .execute()
            )

    # -----------------------------------------------------
    # 備註
    # -----------------------------------------------------

    setting = load_settings(branch)

    if not setting:

        (
            supabase
            .table("daycare_settings")
            .insert(
                {
                    "branch": branch,
                    "notes": DEFAULT_NOTES,
                }
            )
            .execute()
        )


# =========================================================
# 執行初始化
# =========================================================

try:

    initialize_branch_data(
        "澳森",
        DEFAULT_TEACHERS_A,
        DEFAULT_TASKS_A,
    )

    initialize_branch_data(
        "澳森文德",
        DEFAULT_TEACHERS_B,
        DEFAULT_TASKS_B,
    )

except Exception as error:

    st.error(
        "❌ Supabase 資料初始化失敗。"
        "如果剛建立資料表，請確認 service_role 的資料表權限。"
    )

    st.exception(error)
    st.stop()


# =========================================================
# 日期工具
# =========================================================

def is_workday(target_date, tw_holidays):

    return (
        target_date.weekday() < 5
        and target_date not in tw_holidays
    )


def get_adjusted_workday(target_date, tw_holidays):

    if is_workday(target_date, tw_holidays):
        return target_date

    current_date = target_date - datetime.timedelta(days=1)

    while not is_workday(
        current_date,
        tw_holidays,
    ):

        current_date -= datetime.timedelta(days=1)

    return current_date


# =========================================================
# Excel
# =========================================================

def generate_cleaning_excel(
    branch_name,
    year_roc,
    month,
    schedule_df,
    notes_text,
):

    output = io.BytesIO()

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = f"{month}月清潔輪值"

    worksheet.sheet_view.showGridLines = False

    # -----------------------------------------------------
    # 列印設定
    # -----------------------------------------------------

    worksheet.page_setup.orientation = "landscape"

    worksheet.page_setup.paperSize = (
        worksheet.PAPERSIZE_A4
    )

    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.4
    worksheet.page_margins.bottom = 0.4

    # -----------------------------------------------------
    # 樣式
    # -----------------------------------------------------

    thin_gray = Side(
        style="thin",
        color="B7B7B7",
    )

    border = Border(
        left=thin_gray,
        right=thin_gray,
        top=thin_gray,
        bottom=thin_gray,
    )

    title_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="EDEDED",
    )

    date_fill = PatternFill(
        fill_type="solid",
        fgColor="F3F8F4",
    )

    teacher_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF7E6",
    )

    # -----------------------------------------------------
    # 標題
    # -----------------------------------------------------

    worksheet.merge_cells("A1:G1")

    title_cell = worksheet["A1"]

    title_cell.value = (
        f"{branch_name} "
        f"{year_roc} 年 {month} 月清潔輪值紀錄表"
    )

    title_cell.font = Font(
        name="微軟正黑體",
        size=16,
        bold=True,
        color="1F497D",
    )

    title_cell.fill = title_fill

    title_cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 30

    # -----------------------------------------------------
    # 表頭
    # -----------------------------------------------------

    headers = [
        "週次",
        "項目",
        "星期一",
        "星期二",
        "星期三",
        "星期四",
        "星期五",
    ]

    for col_idx, header in enumerate(
        headers,
        start=1,
    ):

        cell = worksheet.cell(
            row=3,
            column=col_idx,
            value=header,
        )

        cell.font = Font(
            name="微軟正黑體",
            size=10,
            bold=True,
        )

        cell.fill = header_fill
        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    worksheet.row_dimensions[3].height = 25

    # -----------------------------------------------------
    # 排班
    # -----------------------------------------------------

    row_num = 4

    day_keys = [
        "一",
        "二",
        "三",
        "四",
        "五",
    ]

    for _, record in schedule_df.iterrows():

        start_row = row_num

        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row + 3,
            end_column=1,
        )

        week_cell = worksheet.cell(
            row=start_row,
            column=1,
            value=record["週次"],
        )

        week_cell.font = Font(
            name="微軟正黑體",
            size=10,
            bold=True,
        )

        week_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        row_types = [
            ("日期", "_0"),
            ("清潔內容", "_1"),
            ("執行老師", "_2"),
            ("簽名", None),
        ]

        for offset, (label, suffix) in enumerate(
            row_types
        ):

            current_row = start_row + offset

            label_cell = worksheet.cell(
                row=current_row,
                column=2,
                value=label,
            )

            label_cell.font = Font(
                name="微軟正黑體",
                size=10,
                bold=True,
            )

            label_cell.border = border

            label_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            if label == "日期":
                label_cell.fill = date_fill

            elif label == "執行老師":
                label_cell.fill = teacher_fill

            for column_index, day_key in enumerate(
                day_keys,
                start=3,
            ):

                if suffix is None:

                    value = ""

                else:

                    value = record[
                        f"{day_key}{suffix}"
                    ]

                cell = worksheet.cell(
                    row=current_row,
                    column=column_index,
                    value=value,
                )

                cell.font = Font(
                    name="微軟正黑體",
                    size=9,
                )

                cell.border = border

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

                if label == "日期":

                    cell.fill = date_fill

                    cell.font = Font(
                        name="微軟正黑體",
                        size=9,
                        bold=True,
                        color="00843D",
                    )

                elif label == "執行老師":

                    cell.fill = teacher_fill

        for r in range(
            start_row,
            start_row + 4,
        ):

            for c in range(
                1,
                8,
            ):

                worksheet.cell(
                    r,
                    c,
                ).border = border

        worksheet.row_dimensions[
            start_row
        ].height = 24

        worksheet.row_dimensions[
            start_row + 1
        ].height = 70

        worksheet.row_dimensions[
            start_row + 2
        ].height = 28

        worksheet.row_dimensions[
            start_row + 3
        ].height = 38

        row_num += 4

    # -----------------------------------------------------
    # 備註
    # -----------------------------------------------------

    notes_row = row_num + 1

    worksheet.merge_cells(
        start_row=notes_row,
        start_column=1,
        end_row=notes_row,
        end_column=7,
    )

    notes_cell = worksheet.cell(
        row=notes_row,
        column=1,
        value=notes_text,
    )

    notes_cell.font = Font(
        name="微軟正黑體",
        size=9,
    )

    notes_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    notes_cell.border = border

    worksheet.row_dimensions[
        notes_row
    ].height = 38

    # -----------------------------------------------------
    # 欄寬
    # -----------------------------------------------------

    worksheet.column_dimensions["A"].width = 9
    worksheet.column_dimensions["B"].width = 11

    for column_letter in [
        "C",
        "D",
        "E",
        "F",
        "G",
    ]:

        worksheet.column_dimensions[
            column_letter
        ].width = 27

    worksheet.freeze_panes = "C4"

    worksheet.print_area = (
        f"A1:G{notes_row}"
    )

    worksheet.oddFooter.center.text = (
        "第 &P 頁／共 &N 頁"
    )

    workbook.save(output)

    output.seek(0)

    return output


# =========================================================
# 預覽 HTML
# =========================================================

def build_preview_html(table_records):

    parts = []

    parts.append(
        '<div class="cleaning-preview-wrapper">'
    )

    parts.append(
        '<table class="cleaning-preview-table">'
    )

    parts.append("<thead>")
    parts.append("<tr>")

    parts.append(
        '<th class="week-col">週次</th>'
    )

    parts.append(
        '<th class="day-col">星期一</th>'
    )

    parts.append(
        '<th class="day-col">星期二</th>'
    )

    parts.append(
        '<th class="day-col">星期三</th>'
    )

    parts.append(
        '<th class="day-col">星期四</th>'
    )

    parts.append(
        '<th class="day-col">星期五</th>'
    )

    parts.append("</tr>")
    parts.append("</thead>")
    parts.append("<tbody>")

    for record in table_records:

        parts.append("<tr>")

        parts.append(
            '<td class="week-col">'
            + html.escape(
                str(record["週次"])
            )
            + "</td>"
        )

        for i in range(5):

            safe_date = html.escape(
                str(record["日期"][i])
            )

            safe_task = html.escape(
                str(record["內容"][i])
            )

            safe_teacher = html.escape(
                str(record["老師"][i])
            )

            parts.append(
                '<td class="day-col">'
            )

            parts.append(
                '<div class="preview-date">'
                + safe_date
                + "</div>"
            )

            parts.append(
                '<div class="preview-task">'
                + safe_task
                + "</div>"
            )

            parts.append(
                '<div class="preview-teacher">'
                + "負責："
                + safe_teacher
                + "</div>"
            )

            parts.append("</td>")

        parts.append("</tr>")

    parts.append("</tbody>")
    parts.append("</table>")
    parts.append("</div>")

    return "".join(parts)


# =========================================================
# 老師管理
# =========================================================

def render_teacher_manager(
    branch,
    prefix,
):

    with st.expander(
        "👥 執行老師名單管理",
        expanded=False,
    ):

        teachers = load_teachers(branch)

        # -------------------------------------------------
        # 新增
        # -------------------------------------------------

        with st.form(
            key=f"teacher_add_form_{prefix}",
            clear_on_submit=True,
        ):

            new_name = st.text_input(
                "新增老師姓名"
            )

            add_clicked = (
                st.form_submit_button(
                    "➕ 新增老師",
                    use_container_width=True,
                )
            )

            if add_clicked:

                new_name = (
                    new_name.strip()
                )

                if not new_name:

                    st.warning(
                        "請輸入老師姓名。"
                    )

                elif any(
                    item["name"] == new_name
                    for item in teachers
                ):

                    st.warning(
                        "此老師已經存在。"
                    )

                else:

                    next_order = (
                        max(
                            [
                                item["sort_order"]
                                for item in teachers
                            ],
                            default=-1,
                        )
                        + 1
                    )

                    (
                        supabase
                        .table("daycare_teachers")
                        .insert(
                            {
                                "branch": branch,
                                "name": new_name,
                                "sort_order": next_order,
                            }
                        )
                        .execute()
                    )

                    st.rerun()

        # -------------------------------------------------
        # 編輯／刪除
        # -------------------------------------------------

        for teacher in teachers:

            edit_col, delete_col = (
                st.columns([9, 1])
            )

            with edit_col:

                edited_name = (
                    st.text_input(
                        "老師姓名",
                        value=teacher["name"],
                        key=(
                            f"teacher_edit_"
                            f"{prefix}_"
                            f"{teacher['id']}"
                        ),
                        label_visibility="collapsed",
                    )
                )

                edited_name = (
                    edited_name.strip()
                )

                if (
                    edited_name
                    and
                    edited_name
                    != teacher["name"]
                ):

                    duplicate = any(
                        other["name"]
                        == edited_name
                        and
                        other["id"]
                        != teacher["id"]
                        for other in teachers
                    )

                    if duplicate:

                        st.warning(
                            f"「{edited_name}」已存在。"
                        )

                    else:

                        (
                            supabase
                            .table("daycare_teachers")
                            .update(
                                {
                                    "name": edited_name
                                }
                            )
                            .eq(
                                "id",
                                teacher["id"],
                            )
                            .execute()
                        )

                        st.rerun()

            with delete_col:

                if st.button(
                    "🗑️",
                    key=(
                        f"teacher_delete_"
                        f"{prefix}_"
                        f"{teacher['id']}"
                    ),
                    help=(
                        f"刪除 {teacher['name']}"
                    ),
                    use_container_width=True,
                ):

                    if len(teachers) <= 1:

                        st.warning(
                            "至少需要保留一位老師。"
                        )

                    else:

                        (
                            supabase
                            .table("daycare_teachers")
                            .delete()
                            .eq(
                                "id",
                                teacher["id"],
                            )
                            .execute()
                        )

                        st.rerun()


# =========================================================
# 清潔工作管理
# =========================================================

def render_task_manager(
    branch,
    prefix,
):

    with st.expander(
        "🧹 清潔工作項目管理",
        expanded=False,
    ):

        tasks = load_tasks(branch)

        # -------------------------------------------------
        # 新增
        # -------------------------------------------------

        with st.form(
            key=f"task_add_form_{prefix}",
            clear_on_submit=True,
        ):

            new_task = st.text_input(
                "新增清潔工作項目"
            )

            add_clicked = (
                st.form_submit_button(
                    "➕ 新增工作項目",
                    use_container_width=True,
                )
            )

            if add_clicked:

                new_task = (
                    new_task.strip()
                )

                if not new_task:

                    st.warning(
                        "請輸入清潔工作內容。"
                    )

                elif any(
                    item["task_name"]
                    == new_task
                    for item in tasks
                ):

                    st.warning(
                        "此清潔工作已經存在。"
                    )

                else:

                    next_order = (
                        max(
                            [
                                item["sort_order"]
                                for item in tasks
                            ],
                            default=-1,
                        )
                        + 1
                    )

                    (
                        supabase
                        .table(
                            "daycare_cleaning_tasks"
                        )
                        .insert(
                            {
                                "branch": branch,
                                "task_name": new_task,
                                "sort_order": next_order,
                            }
                        )
                        .execute()
                    )

                    st.rerun()

        # -------------------------------------------------
        # 編輯／刪除
        # -------------------------------------------------

        for task in tasks:

            edit_col, delete_col = (
                st.columns([9, 1])
            )

            with edit_col:

                edited_task = (
                    st.text_input(
                        "工作項目",
                        value=task["task_name"],
                        key=(
                            f"task_edit_"
                            f"{prefix}_"
                            f"{task['id']}"
                        ),
                        label_visibility="collapsed",
                    )
                )

                edited_task = (
                    edited_task.strip()
                )

                if (
                    edited_task
                    and
                    edited_task
                    != task["task_name"]
                ):

                    duplicate = any(
                        other["task_name"]
                        == edited_task
                        and
                        other["id"]
                        != task["id"]
                        for other in tasks
                    )

                    if duplicate:

                        st.warning(
                            "此清潔工作已存在。"
                        )

                    else:

                        (
                            supabase
                            .table(
                                "daycare_cleaning_tasks"
                            )
                            .update(
                                {
                                    "task_name": (
                                        edited_task
                                    )
                                }
                            )
                            .eq(
                                "id",
                                task["id"],
                            )
                            .execute()
                        )

                        st.rerun()

            with delete_col:

                if st.button(
                    "🗑️",
                    key=(
                        f"task_delete_"
                        f"{prefix}_"
                        f"{task['id']}"
                    ),
                    help="刪除此清潔項目",
                    use_container_width=True,
                ):

                    if len(tasks) <= 1:

                        st.warning(
                            "至少需要保留一個清潔工作項目。"
                        )

                    else:

                        (
                            supabase
                            .table(
                                "daycare_cleaning_tasks"
                            )
                            .delete()
                            .eq(
                                "id",
                                task["id"],
                            )
                            .execute()
                        )

                        st.rerun()


# =========================================================
# 標題
# =========================================================

st.title(
    "🏫 澳森托嬰中心 月清潔輪值表"
)

st.caption(
    "☁️ Supabase 永久儲存版｜"
    "老師、清潔項目、備註與歷史排班皆永久保存"
)


# =========================================================
# Tabs
# =========================================================

tab_a, tab_b = st.tabs(
    [
        "🌳 澳森分園",
        "🌸 澳森文德分園",
    ]
)


# =========================================================
# 分園主畫面
# =========================================================

def render_branch_tab(
    branch_name,
    prefix,
):

    # -----------------------------------------------------
    # 讀取永久資料
    # -----------------------------------------------------

    teacher_rows = (
        load_teachers(
            branch_name
        )
    )

    task_rows = (
        load_tasks(
            branch_name
        )
    )

    setting = (
        load_settings(
            branch_name
        )
    )

    teacher_pool = [
        item["name"]
        for item in teacher_rows
    ]

    current_tasks = [
        item["task_name"]
        for item in task_rows
    ]

    if not teacher_pool:

        teacher_pool = ["主任"]

    if not current_tasks:

        current_tasks = [
            "一般清潔"
        ]

    # =====================================================
    # 基本設定
    # =====================================================

    st.subheader(
        f"📌 {branch_name}｜基本設定"
    )

    col1, col2, col3 = (
        st.columns(
            [1, 1, 1.4]
        )
    )

    with col1:

        year_roc = st.number_input(
            "民國年份",
            min_value=114,
            max_value=125,
            value=115,
            step=1,
            key=f"year_{prefix}",
        )

    with col2:

        month = st.selectbox(
            "月份",
            options=list(
                range(1, 13)
            ),
            index=8,
            key=f"month_{prefix}",
        )

    # 先建立變數
    save_clicked = False

    with col3:

        st.markdown(
            "<div style='height:22px'></div>",
            unsafe_allow_html=True,
        )

        save_clicked = st.button(
            "💾 永久儲存目前排班",
            key=(
                f"save_history_"
                f"{prefix}"
            ),
            use_container_width=True,
        )

    # =====================================================
    # 老師／工作項目管理
    # =====================================================

    manager_left, manager_right = (
        st.columns(2)
    )

    with manager_left:

        render_teacher_manager(
            branch_name,
            prefix,
        )

    with manager_right:

        render_task_manager(
            branch_name,
            prefix,
        )

    st.divider()

    # =====================================================
    # 月曆
    # =====================================================

    year_ad = (
        int(year_roc)
        + 1911
    )

    tw_holidays = (
        holidays.Taiwan(
            years=year_ad
        )
    )

    month_calendar = (
        calendar.monthcalendar(
            year_ad,
            int(month),
        )
    )

    work_weeks = []

    for week in month_calendar:

        weekday_part = (
            week[:5]
        )

        if any(
            day > 0
            for day in weekday_part
        ):

            work_weeks.append(
                weekday_part
            )

    st.subheader(
        f"📅 {branch_name} "
        f"{year_roc} 年 "
        f"{month} 月清潔輪值"
    )

    table_records = []

    task_rotation_index = 0

    week_names = [
        "第一週",
        "第二週",
        "第三週",
        "第四週",
        "第五週",
        "第六週",
    ]

    day_names = [
        "週一",
        "週二",
        "週三",
        "週四",
        "週五",
    ]

    # =====================================================
    # 每週
    # =====================================================

    for week_index, week in enumerate(
        work_weeks
    ):

        week_name = (
            week_names[
                week_index
            ]
        )

        with st.expander(
            f"📌 {week_name}",
            expanded=True,
        ):

            row_dates = []
            row_tasks = []
            row_teachers = []

            columns = (
                st.columns(5)
            )

            # =================================================
            # 每天
            # =================================================

            for (
                day_index,
                day_number,
            ) in enumerate(week):

                with columns[
                    day_index
                ]:

                    # -----------------------------------------
                    # 日期
                    # -----------------------------------------

                    if day_number > 0:

                        target_date = (
                            datetime.date(
                                year_ad,
                                int(month),
                                day_number,
                            )
                        )

                        adjusted_date = (
                            get_adjusted_workday(
                                target_date,
                                tw_holidays,
                            )
                        )

                        date_str = (
                            f"{month}/"
                            f"{day_number}"
                        )

                        if (
                            adjusted_date
                            != target_date
                        ):

                            date_str += (
                                " "
                                "(調日至"
                                f"{adjusted_date.month}/"
                                f"{adjusted_date.day}"
                                ")"
                            )

                    else:

                        date_str = "－"

                    # -----------------------------------------
                    # 星期
                    # -----------------------------------------

                    day_header_html = (
                        '<div class="day-header">'
                        '<span class="day-name">'
                        + html.escape(
                            day_names[
                                day_index
                            ]
                        )
                        + "</span>"
                        '<span class="day-date">'
                        + html.escape(
                            date_str
                        )
                        + "</span>"
                        "</div>"
                    )

                    st.markdown(
                        day_header_html,
                        unsafe_allow_html=True,
                    )

                    # =========================================
                    # 月份外
                    # =========================================

                    if day_number == 0:

                        chosen_task = (
                            st.selectbox(
                                "空白工作",
                                options=["無"],
                                index=0,
                                key=(
                                    f"empty_task_"
                                    f"{prefix}_"
                                    f"{year_roc}_"
                                    f"{month}_"
                                    f"{week_index}_"
                                    f"{day_index}"
                                ),
                                label_visibility="collapsed",
                                disabled=True,
                            )
                        )

                        chosen_teacher = (
                            st.selectbox(
                                "空白老師",
                                options=["無"],
                                index=0,
                                key=(
                                    f"empty_teacher_"
                                    f"{prefix}_"
                                    f"{year_roc}_"
                                    f"{month}_"
                                    f"{week_index}_"
                                    f"{day_index}"
                                ),
                                label_visibility="collapsed",
                                disabled=True,
                            )
                        )

                    # =========================================
                    # 有日期
                    # =========================================

                    else:

                        default_task = (
                            current_tasks[
                                task_rotation_index
                                % len(
                                    current_tasks
                                )
                            ]
                        )

                        # 第二、四週星期二
                        if (
                            day_index == 1
                            and
                            (
                                week_index + 1
                                in [2, 4]
                            )
                        ):

                            matches = [
                                task
                                for task
                                in current_tasks
                                if "清點備品"
                                in task
                            ]

                            if matches:

                                default_task = (
                                    matches[0]
                                )

                        # 星期五
                        elif (
                            day_index == 4
                        ):

                            matches = [
                                task
                                for task
                                in current_tasks
                                if (
                                    "掃地機器人"
                                    in task
                                    or
                                    "戶外掃落葉"
                                    in task
                                    or
                                    "規劃戶外活動"
                                    in task
                                )
                            ]

                            if matches:

                                default_task = (
                                    matches[0]
                                )

                        else:

                            task_rotation_index += 1

                        chosen_task = (
                            st.selectbox(
                                "工作",
                                options=current_tasks,
                                index=(
                                    current_tasks.index(
                                        default_task
                                    )
                                    if
                                    default_task
                                    in current_tasks
                                    else 0
                                ),
                                key=(
                                    f"task_"
                                    f"{prefix}_"
                                    f"{year_roc}_"
                                    f"{month}_"
                                    f"{week_index}_"
                                    f"{day_index}"
                                ),
                                label_visibility="collapsed",
                            )
                        )

                        # -------------------------------------
                        # 老師輪值
                        # -------------------------------------

                        default_teacher_index = (
                            (
                                week_index * 5
                                + day_index
                            )
                            % len(
                                teacher_pool
                            )
                        )

                        default_teacher = (
                            teacher_pool[
                                default_teacher_index
                            ]
                        )

                        chosen_teacher = (
                            st.selectbox(
                                "老師",
                                options=teacher_pool,
                                index=(
                                    teacher_pool.index(
                                        default_teacher
                                    )
                                ),
                                key=(
                                    f"teacher_"
                                    f"{prefix}_"
                                    f"{year_roc}_"
                                    f"{month}_"
                                    f"{week_index}_"
                                    f"{day_index}"
                                ),
                                label_visibility="collapsed",
                            )
                        )

                    row_dates.append(
                        date_str
                    )

                    row_tasks.append(
                        chosen_task
                    )

                    row_teachers.append(
                        chosen_teacher
                    )

            table_records.append(
                {
                    "週次": week_name,
                    "日期": row_dates,
                    "內容": row_tasks,
                    "老師": row_teachers,
                }
            )

    # =====================================================
    # 備註永久儲存
    # =====================================================

    st.markdown(
        "#### 📝 備註"
    )

    current_notes = (
        setting["notes"]
        if setting
        and setting.get("notes")
        else DEFAULT_NOTES
    )

    notes = st.text_area(
        "備註",
        value=current_notes,
        height=70,
        key=(
            f"notes_area_"
            f"{prefix}"
        ),
        label_visibility="collapsed",
    )

    if (
        notes
        != current_notes
    ):

        if setting:

            (
                supabase
                .table(
                    "daycare_settings"
                )
                .update(
                    {
                        "notes": notes,
                        "updated_at": (
                            datetime.datetime.now(
                                datetime.timezone.utc
                            ).isoformat()
                        ),
                    }
                )
                .eq(
                    "id",
                    setting["id"],
                )
                .execute()
            )

        else:

            (
                supabase
                .table(
                    "daycare_settings"
                )
                .insert(
                    {
                        "branch": (
                            branch_name
                        ),
                        "notes": notes,
                    }
                )
                .execute()
            )

    # =====================================================
    # 轉 DataFrame
    # =====================================================

    export_rows = []

    day_keys = [
        "一",
        "二",
        "三",
        "四",
        "五",
    ]

    for record in table_records:

        row_dict = {
            "週次": (
                record["週次"]
            )
        }

        for (
            i,
            day_key,
        ) in enumerate(
            day_keys
        ):

            row_dict[
                f"{day_key}_0"
            ] = record[
                "日期"
            ][i]

            row_dict[
                f"{day_key}_1"
            ] = record[
                "內容"
            ][i]

            row_dict[
                f"{day_key}_2"
            ] = record[
                "老師"
            ][i]

        export_rows.append(
            row_dict
        )

    final_df = (
        pd.DataFrame(
            export_rows
        )
    )

    # =====================================================
    # 永久儲存目前排班
    # =====================================================

    if save_clicked:

        try:

            schedule_json = (
                final_df.to_dict(
                    orient="records"
                )
            )

            title = (
                f"{year_roc}年"
                f"{month}月 "
                "清潔輪值表"
            )

            (
                supabase
                .table(
                    "daycare_schedule_history"
                )
                .insert(
                    {
                        "branch": (
                            branch_name
                        ),
                        "year_roc": (
                            int(year_roc)
                        ),
                        "month": (
                            int(month)
                        ),
                        "title": title,
                        "schedule_data": (
                            schedule_json
                        ),
                        "notes": notes,
                    }
                )
                .execute()
            )

            st.success(
                "✅ 已永久儲存目前排班。"
            )

        except Exception as error:

            st.error(
                "❌ 排班永久儲存失敗。"
            )

            st.exception(error)

    # =====================================================
    # 即時預覽
    # =====================================================

    st.divider()

    st.subheader(
        "👁️ 月曆式即時預覽"
    )

    preview_html = (
        build_preview_html(
            table_records
        )
    )

    st.markdown(
        preview_html,
        unsafe_allow_html=True,
    )

    # =====================================================
    # Excel
    # =====================================================

    st.divider()

    st.subheader(
        f"📥 匯出 "
        f"{branch_name} Excel"
    )

    try:

        excel_data = (
            generate_cleaning_excel(
                branch_name=(
                    branch_name
                ),
                year_roc=(
                    int(year_roc)
                ),
                month=(
                    int(month)
                ),
                schedule_df=(
                    final_df
                ),
                notes_text=notes,
            )
        )

        st.download_button(
            label=(
                "📥 下載【"
                f"{branch_name} "
                f"{year_roc}年"
                f"{month}月"
                "清潔輪值表】"
            ),
            data=excel_data,
            file_name=(
                f"{branch_name}_"
                f"{year_roc}年"
                f"{month}月"
                "清潔輪值紀錄表.xlsx"
            ),
            mime=(
                "application/"
                "vnd.openxmlformats-"
                "officedocument."
                "spreadsheetml.sheet"
            ),
            key=(
                f"excel_download_"
                f"{prefix}_"
                f"{year_roc}_"
                f"{month}"
            ),
            use_container_width=True,
        )

    except Exception as error:

        st.error(
            "❌ Excel 產生失敗。"
        )

        st.exception(error)

    # =====================================================
    # 永久歷史紀錄
    # =====================================================

    st.divider()

    st.subheader(
        f"📂 永久歷史紀錄｜"
        f"{branch_name}"
    )

    try:

        history_list = (
            load_history(
                branch_name
            )
        )

    except Exception as error:

        st.error(
            "❌ 無法讀取歷史紀錄。"
        )

        st.exception(error)

        history_list = []

    if not history_list:

        st.info(
            "目前尚無永久歷史紀錄。"
        )

    else:

        for history in history_list:

            created_at = (
                str(
                    history.get(
                        "created_at",
                        "",
                    )
                )
            )

            display_time = (
                created_at
                .replace(
                    "T",
                    " ",
                )[:16]
            )

            with st.expander(
                (
                    "📜 "
                    f"{history['title']} "
                    f"｜{display_time}"
                ),
                expanded=False,
            ):

                history_notes = (
                    history.get(
                        "notes",
                        "",
                    )
                    or ""
                )

                st.markdown(
                    "**備註：** "
                    + html.escape(
                        history_notes
                    )
                )

                history_data = (
                    history.get(
                        "schedule_data",
                        [],
                    )
                    or []
                )

                history_df = (
                    pd.DataFrame(
                        history_data
                    )
                )

                # -----------------------------------------
                # 歷史預覽
                # -----------------------------------------

                history_records = []

                for (
                    _,
                    history_row,
                ) in history_df.iterrows():

                    history_records.append(
                        {
                            "週次": (
                                history_row[
                                    "週次"
                                ]
                            ),

                            "日期": [
                                history_row[
                                    "一_0"
                                ],
                                history_row[
                                    "二_0"
                                ],
                                history_row[
                                    "三_0"
                                ],
                                history_row[
                                    "四_0"
                                ],
                                history_row[
                                    "五_0"
                                ],
                            ],

                            "內容": [
                                history_row[
                                    "一_1"
                                ],
                                history_row[
                                    "二_1"
                                ],
                                history_row[
                                    "三_1"
                                ],
                                history_row[
                                    "四_1"
                                ],
                                history_row[
                                    "五_1"
                                ],
                            ],

                            "老師": [
                                history_row[
                                    "一_2"
                                ],
                                history_row[
                                    "二_2"
                                ],
                                history_row[
                                    "三_2"
                                ],
                                history_row[
                                    "四_2"
                                ],
                                history_row[
                                    "五_2"
                                ],
                            ],
                        }
                    )

                if history_records:

                    st.markdown(
                        build_preview_html(
                            history_records
                        ),
                        unsafe_allow_html=True,
                    )

                # -----------------------------------------
                # 歷史 Excel
                # -----------------------------------------

                try:

                    history_excel = (
                        generate_cleaning_excel(
                            branch_name=(
                                branch_name
                            ),
                            year_roc=(
                                int(
                                    history[
                                        "year_roc"
                                    ]
                                )
                            ),
                            month=(
                                int(
                                    history[
                                        "month"
                                    ]
                                )
                            ),
                            schedule_df=(
                                history_df
                            ),
                            notes_text=(
                                history_notes
                            ),
                        )
                    )

                    download_col, delete_col = (
                        st.columns(
                            [4, 1]
                        )
                    )

                    with download_col:

                        st.download_button(
                            label=(
                                "📥 下載此歷史版本"
                            ),
                            data=history_excel,
                            file_name=(
                                f"{branch_name}_"
                                f"{history['title']}_"
                                "歷史存檔.xlsx"
                            ),
                            mime=(
                                "application/"
                                "vnd.openxmlformats-"
                                "officedocument."
                                "spreadsheetml.sheet"
                            ),
                            key=(
                                f"history_excel_"
                                f"{prefix}_"
                                f"{history['id']}"
                            ),
                            use_container_width=True,
                        )

                    with delete_col:

                        if st.button(
                            "🗑️ 刪除",
                            key=(
                                f"history_delete_"
                                f"{prefix}_"
                                f"{history['id']}"
                            ),
                            help=(
                                "永久刪除此歷史紀錄"
                            ),
                            use_container_width=True,
                        ):

                            (
                                supabase
                                .table(
                                    "daycare_schedule_history"
                                )
                                .delete()
                                .eq(
                                    "id",
                                    history["id"],
                                )
                                .execute()
                            )

                            st.rerun()

                except Exception as error:

                    st.error(
                        "❌ 此歷史版本 "
                        "Excel 產生失敗。"
                    )

                    st.exception(error)


# =========================================================
# 澳森
# =========================================================

with tab_a:

    render_branch_tab(
        "澳森",
        "A",
    )


# =========================================================
# 澳森文德
# =========================================================

with tab_b:

    render_branch_tab(
        "澳森文德",
        "B",
    )
